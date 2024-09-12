import os.path
import pickle
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import base64
import email
from email.header import decode_header
import openpyxl
from openai import OpenAI
import datetime

# If modifying these SCOPES, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']


# Function to authenticate and connect to Gmail API
def authenticate_gmail():
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)

    # If no valid credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('gmail', 'v1', credentials=creds)
    return service

# Function to get email messages from Gmail inbox
def get_emails(service, query=''):
    # Get list of emails matching the query
    results = service.users().messages().list(userId='me', q=query).execute()
    messages = results.get('messages', [])
    return messages

# Function to fetch and parse a specific email by ID
def fetch_email(service, msg_id):
    msg = service.users().messages().get(userId='me', id=msg_id).execute()
    msg_data = msg['payload']

    subject = ""
    body = ""

    # Decode the subject
    for header in msg_data['headers']:
        if header['name'] == 'Subject':
            subject, encoding = decode_header(header['value'])[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding if encoding else 'utf-8')

    # Extract the message body
    if 'parts' in msg_data:
        for part in msg_data['parts']:
            if part['mimeType'] == 'text/plain':
                body = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8')

    return subject, body

# Updated function to classify and extract details from the body using OpenAI API in a single output
def classify_and_extract_details_from_body(body):
    client = OpenAI(
        api_key=os.getenv("OPENAI_API_KEY")
    )

    # OpenAI prompt to classify the email and extract company and role
    messages = [
        {
            "role": "system",
            "content": "You are a helpful assistant. Classify the email and extract details."
        },
        {
            "role": "user",
            "content": f"""
            Given the following email body:

            {body}

            - Classify the email into one of these categories: 'T' for received, 'F' for rejected, 'I' for moving on to the next phase, or 'None' if none apply.
            - Identify the company mentioned in the email.
            - Identify the role being applied for.

            Return the output as a list in the format: [status, company, role].
            """
        }
    ]

    # Use OpenAI to classify and extract details using chat model
    response = client.chat.completions.create(
        model="gpt-4",
        messages=messages,
        max_tokens=150,
        temperature=0.7
    )

    # Accessing content directly from the message
    output = response.choices[0].message.content.strip().strip("[]").split(", ")

    # Ensure the output is valid
    if len(output) == 3:
        status, company, role = output
        status = status.strip("'\"")  # Remove any surrounding quotes
        company = company.strip("'\"")
        role = role.strip("'\"")
    else:
        status, company, role = None, "Unknown Company", "Unknown Role"

    return status, company, role

# Update or add entry to the Excel file
def update_excel_sheet(company, role, email_datetime, status):
    filename = "application_tracking.xlsx"

    # Load or create Excel file
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Company", "Role", "Date/Time", "Status"])  # Headers
    else:
        ws = wb.active

    # Check if the company and role already exist, if so, update the status
    row_updated = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value == company and row[1].value == role:
            row[2].value = email_datetime
            row[3].value = status
            row_updated = True
            break

    # If no matching row, add a new row
    if not row_updated:
        ws.append([company, role, email_datetime, status])

    wb.save(filename)
    print(f"Updated Excel with: {company}, {role}, {email_datetime}, {status}")

# Run the bot to read emails and update the Excel sheet
def run_email_to_excel_bot():
    # Authenticate and connect to Gmail
    service = authenticate_gmail()

    # Get the current date and the date 10 days ago for the query
    ten_days_ago = (datetime.datetime.now() - datetime.timedelta(days=10)).strftime("%Y/%m/%d")
    query = f"after:{ten_days_ago} subject:application OR subject:status"

    # Get emails from the past 10 days
    emails = get_emails(service, query)

    for email_data in emails:
        # Fetch and parse each email by its ID
        _, body = fetch_email(service, email_data['id'])

        # Use the updated function to classify and extract details
        status, company, role = classify_and_extract_details_from_body(body)

        if status:
            # Extract email date/time
            email_datetime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Update Excel sheet
            update_excel_sheet(company, role, email_datetime, status)

run_email_to_excel_bot()
